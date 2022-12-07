import asyncio
from datetime import datetime

import openpyxl
import requests
from aiohttp import ClientSession
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from data import Month_names
from data import headers, marks


class RuobrException(Exception):
    class NoDateException(Exception):
        pass

    class DateException(Exception):
        pass

    class AuthorizationException(Exception):
        pass

    class Ruobr503Exception(Exception):
        pass


class Ruobr:

    def __init__(self, username, password):
        self.__username, self.__password = username, password
        self.__cookie = self.__cookies()

        if len(self.__cookie) != 2:
            raise RuobrException.AuthorizationException('Логин или пароль неверные')

    def __cookies(self):
        session = requests.Session()
        session.get('https://cabinet.ruobr.ru/login/', headers=headers)
        try:
            data = {'username': f'{self.__username}', 'password': f'{self.__password}',
                    'csrfmiddlewaretoken': dict(session.cookies)['csrftoken']}
        except KeyError:
            raise RuobrException.Ruobr503Exception('Ошибка на сайте Cabinet Ruobr, попробуйте позже')
        session.post('https://cabinet.ruobr.ru/login/', headers=headers, data=data)
        return dict(session.cookies)

    def __get_esimation(self):

        async def __taskmaster():
            pagination = await __pagination()
            tasks = []
            urls = (f'https://cabinet.ruobr.ru//student/progress/?page={num}' for num in range(1, pagination))

            async with ClientSession() as session:
                for url in urls:
                    tasks.append(asyncio.create_task(__fetch(url, session)))
                    await asyncio.sleep(0.02)

                return await asyncio.gather(*tasks)

        async def __fetch(url, session):
            async with session.get(url, cookies=self.__cookie) as response:
                return await response.text()

        async def __pagination():
            async with ClientSession() as session:
                async with session.get('https://cabinet.ruobr.ru//student/progress/?page=1',
                                       cookies=self.__cookie) as response:
                    soup = BeautifulSoup(await response.text(), 'html.parser')
                    pagination = int([link.get('href') for link in soup.find_all('a')][-2].split('=')[1]) + 1
                    return pagination

        async def __handler():
            all_marks = []
            for request in await __taskmaster():
                soup = BeautifulSoup(request, 'html.parser')
                data = list(map(str, soup.find_all('tr')))
                for t in data:
                    cleantext = BeautifulSoup(t, 'html.parser').text.split('\n')[3:6]
                    if cleantext != ['Дата', 'Дисциплина', 'Отметка']:
                        all_marks.append(cleantext)

            '''Изменение оценок на числа'''
            for i, ell in enumerate(all_marks):
                all_marks[i][2] = marks[f'{all_marks[i][2]}']

            '''Изменение дат на реверсивные им'''
            for i, ell in enumerate(all_marks):
                newdata = all_marks[i][0].split('.')
                newdata = '.'.join(newdata[::-1])
                all_marks[i][0] = newdata

            return all_marks

        return asyncio.run(__handler())

    def marks(self, date=None):

        def __student_year(date):
            start = datetime.strptime(f'{date}0901', "%Y%m%d").date()
            dates = [(start + relativedelta(months=date)).strftime('%Y.%m') for date in range(1, 9)]
            dates.insert(0, start.strftime('%Y.%m'))
            return dates

        all_marks = self.__get_esimation()

        match date:
            case 'year':
                date = datetime.today().strftime('%Y')
            case 'month':
                date = datetime.today().strftime('%Y.%m')
            case 'day':
                date = datetime.today().strftime('%Y.%m.%d')
            case None:
                return 'Дата неуказана'

        match len(date):
            case 4:
                sorted_marks = [ell for ell in all_marks if ell[0][:7] in __student_year(date)]
                sorted_marks.sort()
            case 7:
                sorted_marks = [ell for ell in all_marks if ell[0][:7] in date]
                sorted_marks.sort()
            case 10:
                sorted_marks = [ell for ell in all_marks if ell[0] in date]
                sorted_marks.sort()
            case _:
                return 'Дата указана неверно'

        return sorted_marks

    def year_estimnation(self, date):
        if date == 'year': date = datetime.today().strftime('%Y')
        all_marks = self.marks(date)
        if len(date) == 4:
            lessons = {lesson[1] for lesson in all_marks}
            year_marks = [f' - {l} - {"%.2f" % (sum(lst := [ell[2] for ell in all_marks if ell[1] == l]) / len(lst))}'
                          for l in
                          lessons]
            year_marks.sort()
            if year_marks:
                result = f'Средний балл за {date} год\n\n' + '\n'.join(year_marks)
            else:
                result = f'Оценок за {date} нет!'
            return result

    def month_estimnation(self, date):
        if date == 'month': date = datetime.today().strftime('%Y.%m')
        all_marks = self.marks(date)
        if len(date) == 7:
            lessons = {lesson[1] for lesson in all_marks}
            year_marks = [f' - {l} - {"%.2f" % (sum(lst := [ell[2] for ell in all_marks if ell[1] == l]) / len(lst))}'
                          for l in
                          lessons]
            year_marks.sort()
            if year_marks:
                result = f'Средний балл за {Month_names[int(date[5:])]} {date[:4]} года\n\n' + '\n'.join(year_marks)
            else:
                result = f'Оценок за {date} нет!'
            return result

    def day_estimnation(self, date):
        if date == 'day': date = datetime.today().strftime('%Y.%m.%d')
        all_marks = self.marks(date)
        if len(date) == 10:
            lessons = {lesson[1] for lesson in all_marks}
            year_marks = [f'{l} - {[ell[2] for ell in all_marks if ell[1] == l]}' for l in lessons]
            year_marks.sort()
            if year_marks:
                result = f'Оценки за {date}\n\n' + '\n'.join(year_marks)
            else:
                result = f'Оценок за {date} нет!'
            return result

    def ToExcel(self, filename: str = 'file'):
        all_marks = self.marks('year')

        def __split(array: list):
            date = {i[0] for i in array}
            date = list(date)
            date.sort()

            subjects = {i[1] for i in array}
            subjects = list(subjects)
            subjects.sort()

            marks = [i[2] for i in array]

            return date, subjects, marks

        def __demarks(array: list):
            mas = [i[0] for i in array]
            for i, ell in enumerate(mas):
                if mas.count(ell) > 1:
                    indexes = [i for i, ltr in enumerate(mas) if ltr == ell]
                    if indexes:
                        marks_from_index = [array[mr][2] for mr in indexes]
                        array[indexes[0]][2] = str(marks_from_index)
                        indexes = indexes[::-1]
                        del indexes[-1]
                        for i in indexes: array.pop(i)
                        for i in indexes: mas.pop(i)
            return array

        def __style():
            def __create_styles():
                """СТИЛЬ ДЛЯ ДВОЕК"""
                ns = NamedStyle(name='BAD')
                ns.fill = PatternFill(fgColor='ff8e7a', fill_type='solid')
                ns.font = Font(color='1f0800', name='Bahnscrift', bold=True)
                ns.alignment = Alignment(horizontal='center', vertical='center')
                wb.add_named_style(ns)

                """СТИЛЬ ДЛЯ ТРОЕК"""
                ns = NamedStyle(name='NORMAL')
                ns.fill = PatternFill(fgColor='ffd500', fill_type='solid')
                ns.alignment = Alignment(horizontal='center', vertical='center')
                ns.font = Font(color='1f0800', name='Bahnscrift', bold=True)
                wb.add_named_style(ns)

                """ПАЦИФИЗМ"""
                ns = NamedStyle(name='PACIFIC')
                ns.fill = PatternFill(fgColor='808080', fill_type='solid')
                ns.alignment = Alignment(horizontal='center', vertical='center')
                wb.add_named_style(ns)

            def __recolor_cell(badly):

                twos, threes = badly

                for column, row in twos:
                    sheet.cell(column=column, row=row).style = "BAD"

                for column, row in threes:
                    sheet.cell(column=column, row=row).style = "NORMAL"

            def __set_to_cells(cell_range):
                for row in sheet[cell_range]:
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            def __size():
                for i in range(2, sheet.max_column + 1):
                    sheet.column_dimensions[f'{get_column_letter(i)}'].width = 5

            __create_styles()
            __recolor_cell(indexes_from_bad_marks)
            __set_to_cells(f'A1:AO{sheet.max_row}')
            __size()

        def __fill_table():

            data = __split(all_marks)
            dates = data[0]
            subjects = data[1]

            row_index = 0

            twos = []
            threes = []

            for i in range(1, 13):

                the_month = [y for y in dates if int(y[5:7]) == i]
                if the_month:
                    month_to_table = [i[5:] for i in the_month]
                    sheet.append([Month_names[int(the_month[0][5:7])]] + month_to_table)
                    row_index += 1
                    for subject in subjects:
                        row_index += 1
                        row = ['' for _ in range(len(the_month))]
                        marks_from_month = [y for y in all_marks if int(y[0][5:7]) == i]
                        marks_from_month_and_subject = [ellements for ellements in marks_from_month if
                                                        subject == ellements[1]]
                        marks_from_month_and_subject = __demarks(marks_from_month_and_subject)

                        for index, mn in enumerate(the_month):
                            for mr in marks_from_month_and_subject:
                                if mn == mr[0]:
                                    row[index] = mr[2]

                        if 2 in row:
                            column_indexes = [i + 2 for i, x in enumerate(row) if x == 2]
                            for column_index in column_indexes:
                                twos.append([column_index, row_index])

                        if 3 in row:
                            column_indexes = [i + 2 for i, x in enumerate(row) if x == 3]
                            for column_index in column_indexes:
                                threes.append([column_index, row_index])
                        sheet.append([subject] + row)

                    row_index += 2

                    sheet.append([''])
                    sheet.append([''])

            return twos, threes

        """ПОДГОТОВКА ДОКУМЕНТА"""
        wb = openpyxl.Workbook()
        title = 'Оценки'
        wb.create_sheet(title=title, index=0)
        del wb['Sheet']
        sheet = wb[title]
        sheet.column_dimensions['A'].width = 100

        indexes_from_bad_marks = __fill_table()
        __style()

        wb.save(f'{filename}.xlsx')
